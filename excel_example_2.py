import openpyxl, os

wb = openpyxl.Workbook()
wb
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value

sheet['A1'].value == None

sheet['A1'].value = 42
sheet['A2'].value = 'Hello'

os.chdir('')
wb.save('example.xlsx')

sheet2 = wb.create_sheet()
wb.get_sheet_names()

sheet2.title

sheet2.title = 'My new Sheet Name"

wb.get_sheet_names()

wb.save('example2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')

wb.save('example3.xlsx')
